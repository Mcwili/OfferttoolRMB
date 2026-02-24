"""
Excel-Parser
Verwendet openpyxl und pandas für Tabellenextraktion
"""

from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import pandas as pd
from io import BytesIO
from datetime import datetime
import re
import base64
import logging
from PIL import Image
from app.models.project import ProjectFile

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser für Excel-Dateien (.xlsx, .xls)"""
    
    async def parse(self, file_content: bytes, file_obj: ProjectFile) -> Dict[str, Any]:
        """
        Extrahiert Daten aus Excel-Datei
        Returns: Dict mit extrahierten Entitäten
        """
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        
        extracted_data = {
            "raeume": [],
            "anlagen": [],
            "geraete": [],
            "anforderungen": [],
            "termine": [],
            "leistungen": [],
            "raw_tables": [],
            "images": [],  # Neue Sektion für Bilder
            "full_text": [],  # Unstrukturierter Text
            "metadata": {
                "sheets": [],
                "total_sheets": len(workbook.sheetnames),
                "file_name": file_obj.original_filename
            }
        }
        
        # Quelle-Info für alle extrahierten Daten
        source_info = {
            "datei": file_obj.original_filename,
            "datei_id": file_obj.id,
            "upload_am": file_obj.upload_date.isoformat() if file_obj.upload_date else None
        }
        
        # Alle Blätter durchgehen
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Metadaten für jedes Blatt sammeln
            sheet_metadata = {
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "has_formulas": False,
                "column_headers": []
            }
            
            # Tabellentyp erkennen
            table_type = self._detect_table_type(sheet_name, sheet)
            
            # Strukturierte Extraktion für erkannte Typen
            if table_type == "raumliste":
                raeume = self._extract_raeume(sheet, source_info)
                extracted_data["raeume"].extend(raeume)
            elif table_type == "geraeteliste":
                geraete = self._extract_geraete(sheet, source_info)
                extracted_data["geraete"].extend(geraete)
            elif table_type == "anlagenliste":
                anlagen = self._extract_anlagen(sheet, source_info)
                extracted_data["anlagen"].extend(anlagen)
            elif table_type == "terminplan":
                termine = self._extract_termine(sheet, source_info)
                extracted_data["termine"].extend(termine)
            elif table_type == "leistungsverzeichnis":
                leistungen = self._extract_leistungen(sheet, source_info)
                extracted_data["leistungen"].extend(leistungen)
            
            # IMMER: Rohdaten extrahieren (auch wenn Typ nicht erkannt wurde)
            raw_table = self._extract_raw_table(sheet, sheet_name, source_info)
            if raw_table:
                extracted_data["raw_tables"].append(raw_table)
            
            # Spaltenüberschriften extrahieren
            header_row = self._find_header_row(sheet)
            if header_row:
                headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
                sheet_metadata["column_headers"] = [str(h) if h else "" for h in headers]
            
            # Prüfen auf Formeln
            for row in sheet.iter_rows(min_row=1, max_row=min(100, sheet.max_row)):
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        sheet_metadata["has_formulas"] = True
                        break
                if sheet_metadata["has_formulas"]:
                    break
            
            extracted_data["metadata"]["sheets"].append(sheet_metadata)
            
            # Bilder aus diesem Blatt extrahieren (mit Timeout-Schutz)
            try:
                sheet_images = await self._extract_images_from_sheet(sheet, sheet_name, source_info, file_obj)
                extracted_data["images"].extend(sheet_images)
            except Exception as e:
                # Bilder-Extraktion sollte nicht die gesamte Extraktion blockieren
                logger.warning(f"Fehler bei Bild-Extraktion aus Blatt {sheet_name}: {e}")
                pass
            
            # Unstrukturierten Text aus diesem Blatt extrahieren
            unstructured_text = self._extract_unstructured_text(sheet, sheet_name, source_info)
            extracted_data["full_text"].extend(unstructured_text)
        
        return extracted_data
    
    def _detect_table_type(self, sheet_name: str, sheet) -> Optional[str]:
        """Erkennt den Tabellentyp eines Blattes"""
        name_lower = sheet_name.lower()
        
        # Raumliste
        if any(kw in name_lower for kw in ["raum", "room", "raeume", "räume"]):
            return "raumliste"
        
        # Geräteliste
        if any(kw in name_lower for kw in ["geraet", "equipment", "gerät"]):
            return "geraeteliste"
        
        # Anlagenliste
        if any(kw in name_lower for kw in ["anlage", "anlagen", "system"]):
            return "anlagenliste"
        
        # Terminplan
        if any(kw in name_lower for kw in ["termin", "zeitplan", "schedule", "gantt", "meilenstein"]):
            return "terminplan"
        
        # Leistungsverzeichnis
        if any(kw in name_lower for kw in ["leistung", "lv", "leistungsverzeichnis", "position"]):
            return "leistungsverzeichnis"
        
        # Anhand der ersten Zeilen erkennen
        first_rows = list(sheet.iter_rows(min_row=1, max_row=5, values_only=True))
        header_text = " ".join([str(cell) for row in first_rows for cell in row if cell]).lower()
        
        if any(kw in header_text for kw in ["raum", "room", "fläche", "flaeche"]):
            return "raumliste"
        if any(kw in header_text for kw in ["geraet", "equipment", "typ", "leistung"]):
            return "geraeteliste"
        if any(kw in header_text for kw in ["anlage", "system", "lüftung", "heizung"]):
            return "anlagenliste"
        if any(kw in header_text for kw in ["datum", "termin", "deadline", "abgabe"]):
            return "terminplan"
        
        return None
    
    def _find_header_row(self, sheet, max_search_rows: int = 20) -> Optional[int]:
        """Findet die Header-Zeile einer Tabelle"""
        # Erweiterte Suche: Prüfe mehr Zeilen und akzeptiere auch einzelne Text-Spalten
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_search_rows, values_only=True), 1):
            if row and any(cell and str(cell).strip() for cell in row):
                # Prüfe ob es wie ein Header aussieht (mindestens eine Text-Spalte)
                text_cells = [cell for cell in row if cell and str(cell).strip()]
                if len(text_cells) >= 1:  # Reduziert von 2 auf 1
                    return row_idx
        # Fallback: Wenn keine Header gefunden, verwende erste Zeile
        if sheet.max_row >= 1:
            return 1
        return None
    
    def _extract_raeume(self, sheet, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert Räume aus einem Blatt"""
        raeume = []
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            return raeume
        
        # Spalten-Indizes finden
        headers = [cell.value for cell in sheet[header_row]]
        column_map = {}
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                # Raum-ID/Name
                if any(kw in header_lower for kw in ["raum", "room", "nummer", "nr", "bezeichnung"]):
                    column_map["raum"] = idx
                # Fläche
                elif any(kw in header_lower for kw in ["flaeche", "fläche", "m²", "m2", "area"]):
                    column_map["flaeche"] = idx
                # Volumen
                elif any(kw in header_lower for kw in ["volumen", "volume", "m³", "m3"]):
                    column_map["volumen"] = idx
                # Höhe
                elif any(kw in header_lower for kw in ["hoehe", "höhe", "höhe m", "height"]):
                    column_map["hoehe"] = idx
                # Nutzung
                elif any(kw in header_lower for kw in ["nutzung", "nutzungsart", "art", "typ", "usage"]):
                    column_map["nutzung"] = idx
                # Geschoss
                elif any(kw in header_lower for kw in ["geschoss", "etage", "floor", "level"]):
                    column_map["geschoss"] = idx
                # Zone
                elif any(kw in header_lower for kw in ["zone", "bereich", "area"]):
                    column_map["zone"] = idx
        
        if "raum" not in column_map:
            return raeume
        
        # Datenzeilen extrahieren (mit Limit für Performance)
        MAX_ROWS_TO_EXTRACT = 10000
        max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
        for row_idx in range(header_row + 1, max_row + 1):
            row = sheet[row_idx]
            raum_value = row[column_map["raum"]].value if column_map["raum"] < len(row) else None
            
            if not raum_value:
                continue
            
            raum_name = str(raum_value).strip()
            if not raum_name or raum_name.lower() in ["", "gesamt", "summe", "total"]:
                continue
            
            # Werte extrahieren
            flaeche = self._extract_number(row, column_map.get("flaeche"))
            volumen = self._extract_number(row, column_map.get("volumen"))
            hoehe = self._extract_number(row, column_map.get("hoehe"))
            nutzung = self._extract_text(row, column_map.get("nutzung"))
            geschoss = self._extract_text(row, column_map.get("geschoss"))
            zone = self._extract_text(row, column_map.get("zone"))
            
            raum = {
                "id": f"Raum_{raum_name.replace(' ', '_')}",
                "name": raum_name,
                "nummer": raum_name,
                "flaeche_m2": flaeche,
                "volumen_m3": volumen,
                "hoehe_m": hoehe,
                "nutzungsart": nutzung,
                "geschoss": geschoss,
                "zone": zone,
                "quelle": {
                    **source_info,
                    "blatt": sheet.title,
                    "zeile": row_idx
                }
            }
            raeume.append(raum)
        
        return raeume
    
    def _extract_number(self, row, col_idx: Optional[int]) -> Optional[float]:
        """Extrahiert eine Zahl aus einer Zelle"""
        if col_idx is None or col_idx >= len(row):
            return None
        value = row[col_idx].value
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Versuche Zahl aus String zu extrahieren
            match = re.search(r'[\d,\.]+', value.replace(',', '.'))
            if match:
                try:
                    return float(match.group())
                except ValueError:
                    pass
        return None
    
    def _extract_text(self, row, col_idx: Optional[int]) -> Optional[str]:
        """Extrahiert Text aus einer Zelle"""
        if col_idx is None or col_idx >= len(row):
            return None
        value = row[col_idx].value
        if value is None:
            return None
        return str(value).strip() if value else None
    
    def _extract_geraete(self, sheet, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert Geräte aus einem Blatt"""
        geraete = []
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            return geraete
        
        headers = [cell.value for cell in sheet[header_row]]
        column_map = {}
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if any(kw in header_lower for kw in ["geraet", "equipment", "gerät", "bezeichnung", "name"]):
                    column_map["geraet"] = idx
                elif any(kw in header_lower for kw in ["typ", "type", "art"]):
                    column_map["typ"] = idx
                elif any(kw in header_lower for kw in ["leistung", "power", "kw", "leistung kw"]):
                    column_map["leistung"] = idx
                elif any(kw in header_lower for kw in ["anlage", "system"]):
                    column_map["anlage"] = idx
                elif any(kw in header_lower for kw in ["raum", "room"]):
                    column_map["raum"] = idx
        
        if "geraet" not in column_map and "typ" not in column_map:
            return geraete
        
        # Datenzeilen extrahieren (mit Limit für Performance)
        MAX_ROWS_TO_EXTRACT = 10000
        max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
        for row_idx in range(header_row + 1, max_row + 1):
            row = sheet[row_idx]
            geraet_name = self._extract_text(row, column_map.get("geraet")) or self._extract_text(row, column_map.get("typ"))
            
            if not geraet_name:
                continue
            
            typ = self._extract_text(row, column_map.get("typ")) or geraet_name
            leistung = self._extract_number(row, column_map.get("leistung"))
            anlage = self._extract_text(row, column_map.get("anlage"))
            raum = self._extract_text(row, column_map.get("raum"))
            
            geraet = {
                "id": f"GER_{geraet_name.replace(' ', '_')}_{row_idx}",
                "typ": typ,
                "name": geraet_name,
                "leistung_kw": leistung,
                "zugehoerige_anlage": anlage,
                "zugehoeriger_raum": raum,
                "quelle": {
                    **source_info,
                    "blatt": sheet.title,
                    "zeile": row_idx
                }
            }
            geraete.append(geraet)
        
        return geraete
    
    def _extract_anlagen(self, sheet, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert Anlagen aus einem Blatt"""
        anlagen = []
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            return anlagen
        
        headers = [cell.value for cell in sheet[header_row]]
        column_map = {}
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if any(kw in header_lower for kw in ["anlage", "system", "bezeichnung"]):
                    column_map["anlage"] = idx
                elif any(kw in header_lower for kw in ["typ", "type", "art"]):
                    column_map["typ"] = idx
                elif any(kw in header_lower for kw in ["leistung", "power", "kw"]):
                    column_map["leistung"] = idx
                elif any(kw in header_lower for kw in ["volumenstrom", "m3/h", "m³/h", "luftmenge"]):
                    column_map["volumenstrom"] = idx
        
        if "anlage" not in column_map and "typ" not in column_map:
            return anlagen
        
        # Datenzeilen extrahieren (mit Limit für Performance)
        MAX_ROWS_TO_EXTRACT = 10000
        max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
        for row_idx in range(header_row + 1, max_row + 1):
            row = sheet[row_idx]
            anlage_name = self._extract_text(row, column_map.get("anlage")) or self._extract_text(row, column_map.get("typ"))
            
            if not anlage_name:
                continue
            
            typ = self._extract_text(row, column_map.get("typ")) or anlage_name
            leistung = self._extract_number(row, column_map.get("leistung"))
            volumenstrom = self._extract_number(row, column_map.get("volumenstrom"))
            
            anlage = {
                "id": f"ANL_{anlage_name.replace(' ', '_')}_{row_idx}",
                "typ": typ,
                "name": anlage_name,
                "leistung_kw": leistung,
                "leistung_m3_h": volumenstrom,
                "quelle": {
                    **source_info,
                    "blatt": sheet.title,
                    "zeile": row_idx
                }
            }
            anlagen.append(anlage)
        
        return anlagen
    
    def _extract_termine(self, sheet, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert Termine aus einem Blatt"""
        termine = []
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            return termine
        
        headers = [cell.value for cell in sheet[header_row]]
        column_map = {}
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if any(kw in header_lower for kw in ["beschreibung", "bezeichnung", "aufgabe", "task"]):
                    column_map["beschreibung"] = idx
                elif any(kw in header_lower for kw in ["datum", "termin", "date", "deadline", "abgabe"]):
                    column_map["datum"] = idx
                elif any(kw in header_lower for kw in ["kategorie", "art", "typ"]):
                    column_map["kategorie"] = idx
                elif any(kw in header_lower for kw in ["phase", "sia"]):
                    column_map["phase"] = idx
        
        if "beschreibung" not in column_map:
            return termine
        
        # Datenzeilen extrahieren (mit Limit für Performance)
        MAX_ROWS_TO_EXTRACT = 10000
        max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
        for row_idx in range(header_row + 1, max_row + 1):
            row = sheet[row_idx]
            beschreibung = self._extract_text(row, column_map.get("beschreibung"))
            
            if not beschreibung:
                continue
            
            datum_value = row[column_map["datum"]].value if column_map.get("datum") is not None and column_map["datum"] < len(row) else None
            termin_datum = None
            if datum_value:
                if isinstance(datum_value, datetime):
                    termin_datum = datum_value.isoformat()
                elif isinstance(datum_value, str):
                    termin_datum = datum_value
                else:
                    termin_datum = str(datum_value)
            
            kategorie = self._extract_text(row, column_map.get("kategorie"))
            phase = self._extract_text(row, column_map.get("phase"))
            
            termin = {
                "id": f"TERM_{row_idx:04d}",
                "beschreibung": beschreibung,
                "termin_datum": termin_datum,
                "kategorie": kategorie or "Meilenstein",
                "sia_phase": phase,
                "quelle": {
                    **source_info,
                    "blatt": sheet.title,
                    "zeile": row_idx
                }
            }
            termine.append(termin)
        
        return termine
    
    def _extract_leistungen(self, sheet, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert Leistungen aus einem Blatt"""
        leistungen = []
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            return leistungen
        
        headers = [cell.value for cell in sheet[header_row]]
        column_map = {}
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if any(kw in header_lower for kw in ["beschreibung", "bezeichnung", "leistung", "position"]):
                    column_map["beschreibung"] = idx
                elif any(kw in header_lower for kw in ["einheit", "unit"]):
                    column_map["einheit"] = idx
                elif any(kw in header_lower for kw in ["menge", "anzahl", "quantity"]):
                    column_map["menge"] = idx
                elif any(kw in header_lower for kw in ["phase", "sia"]):
                    column_map["phase"] = idx
                elif any(kw in header_lower for kw in ["kategorie", "art"]):
                    column_map["kategorie"] = idx
        
        if "beschreibung" not in column_map:
            return leistungen
        
        # Datenzeilen extrahieren (mit Limit für Performance)
        MAX_ROWS_TO_EXTRACT = 10000
        max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
        for row_idx in range(header_row + 1, max_row + 1):
            row = sheet[row_idx]
            beschreibung = self._extract_text(row, column_map.get("beschreibung"))
            
            if not beschreibung:
                continue
            
            einheit = self._extract_text(row, column_map.get("einheit"))
            menge = self._extract_number(row, column_map.get("menge"))
            phase = self._extract_text(row, column_map.get("phase"))
            kategorie = self._extract_text(row, column_map.get("kategorie"))
            
            leistung = {
                "id": f"LEIST_{row_idx:04d}",
                "beschreibung": beschreibung,
                "einheit": einheit,
                "menge": menge,
                "sia_phase": phase,
                "kategorie": kategorie,
                "quelle": {
                    **source_info,
                    "blatt": sheet.title,
                    "zeile": row_idx
                }
            }
            leistungen.append(leistung)
        
        return leistungen
    
    async def _extract_images_from_sheet(self, sheet, sheet_name: str, source_info: Dict[str, Any], file_obj: ProjectFile) -> List[Dict[str, Any]]:
        """Extrahiert Bilder und Grafiken aus einem Excel-Blatt (mit Limit für Performance)"""
        images = []
        MAX_IMAGES_PER_SHEET = 50  # Begrenze Anzahl der Bilder pro Blatt
        
        # Bilder aus openpyxl extrahieren
        if hasattr(sheet, '_images'):
            for img_idx, img in enumerate(sheet._images[:MAX_IMAGES_PER_SHEET]):  # Limit hier setzen
                try:
                    image_blob = img._data()
                    
                    # Bildformat erkennen
                    image_format = None
                    if image_blob[:2] == b'\xff\xd8':
                        image_format = 'jpeg'
                    elif image_blob[:8] == b'\x89PNG\r\n\x1a\n':
                        image_format = 'png'
                    elif image_blob[:6] in [b'GIF87a', b'GIF89a']:
                        image_format = 'gif'
                    elif image_blob[:2] == b'BM':
                        image_format = 'bmp'
                    else:
                        image_format = 'unknown'
                    
                    # Position im Sheet
                    anchor = img.anchor
                    cell_ref = None
                    if hasattr(anchor, '_from'):
                        cell_ref = f"{anchor._from.col}{anchor._from.row}"
                    
                    # Bildgröße
                    width = img.width if hasattr(img, 'width') else None
                    height = img.height if hasattr(img, 'height') else None
                    
                    # Bild als Base64 kodieren (nur Metadaten, nicht vollständiges Bild)
                    # Für große Bilder nur einen Teil speichern
                    image_base64_preview = base64.b64encode(image_blob[:5000]).decode('utf-8') if len(image_blob) > 5000 else base64.b64encode(image_blob).decode('utf-8')
                    
                    # Versuche Bild zu öffnen für weitere Metadaten
                    try:
                        pil_image = Image.open(BytesIO(image_blob))
                        image_width, image_height = pil_image.size
                        image_mode = pil_image.mode
                        
                        # OCR auf Bild anwenden (nur bei kleinen Bildern, um Performance zu schonen)
                        ocr_text = None
                        if image_width * image_height < 1000000:  # Nur bei Bildern < 1MP
                            try:
                                import pytesseract
                                from app.core.config import settings
                                ocr_text = pytesseract.image_to_string(
                                    pil_image,
                                    lang=settings.OCR_LANGUAGE if hasattr(settings, 'OCR_LANGUAGE') else 'deu',
                                    config='--psm 6'
                                ).strip()
                            except Exception:
                                pass
                        
                        image_data = {
                            "id": f"IMG_{file_obj.id}_{sheet_name}_{img_idx}",
                            "index": img_idx,
                            "sheet": sheet_name,
                            "cell_reference": cell_ref,
                            "format": image_format,
                            "width_px": image_width,
                            "height_px": image_height,
                            "width_excel": width,
                            "height_excel": height,
                            "mode": image_mode,
                            "size_bytes": len(image_blob),
                            "data_base64": image_base64_preview[:500] + "..." if len(image_base64_preview) > 500 else image_base64_preview,
                            "has_text": bool(ocr_text),
                            "ocr_text": ocr_text if ocr_text else None,
                            "quelle": {
                                **source_info,
                                "blatt": sheet_name,
                                "typ": "image",
                                "position": img_idx
                            }
                        }
                    except Exception as e:
                        image_data = {
                            "id": f"IMG_{file_obj.id}_{sheet_name}_{img_idx}",
                            "index": img_idx,
                            "sheet": sheet_name,
                            "cell_reference": cell_ref,
                            "format": image_format,
                            "width_excel": width,
                            "height_excel": height,
                            "size_bytes": len(image_blob),
                            "error": str(e),
                            "quelle": {
                                **source_info,
                                "blatt": sheet_name,
                                "typ": "image",
                                "position": img_idx
                            }
                        }
                    
                    images.append(image_data)
                    
                except Exception as e:
                    logger.warning(f"Fehler beim Extrahieren von Bild {img_idx} aus Blatt {sheet_name}: {e}")
                    continue
        
        return images
    
    def _extract_raw_table(self, sheet, sheet_name: str, source_info: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Extrahiert alle Tabellendaten als Rohdaten (auch wenn Typ nicht erkannt wurde)"""
        if sheet.max_row < 1:  # Mindestens eine Zeile
            return None
        
        header_row = self._find_header_row(sheet)
        if not header_row:
            # Fallback: Verwende erste Zeile als Header
            header_row = 1
        
        # Header extrahieren
        header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        # Erweitere Header-Liste auf maximale Spaltenanzahl
        max_col = sheet.max_column
        headers = []
        for i in range(max_col):
            if i < len(header_cells):
                cell_value = header_cells[i]
                if cell_value and str(cell_value).strip():
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Spalte_{i+1}")
            else:
                headers.append(f"Spalte_{i+1}")
        
        # Datenzeilen extrahieren (ab Zeile nach Header)
        # BEGRENZUNG: Maximal 10.000 Zeilen extrahieren, um Performance-Probleme zu vermeiden
        MAX_ROWS_TO_EXTRACT = 10000
        rows_data = []
        start_data_row = header_row + 1 if header_row < sheet.max_row else header_row
        max_row_to_process = min(sheet.max_row, start_data_row + MAX_ROWS_TO_EXTRACT)
        
        empty_rows_count = 0
        MAX_CONSECUTIVE_EMPTY_ROWS = 100  # Stoppe, wenn 100 leere Zeilen hintereinander kommen
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_data_row, max_row=max_row_to_process, values_only=True), start=start_data_row):
            # Überspringe komplett leere Zeilen
            if not any(cell for cell in row):
                empty_rows_count += 1
                if empty_rows_count >= MAX_CONSECUTIVE_EMPTY_ROWS:
                    # Stoppe Extraktion, wenn zu viele leere Zeilen hintereinander kommen
                    break
                continue
            else:
                empty_rows_count = 0  # Reset counter wenn Zeile Daten hat
            
            row_dict = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers):
                    # Konvertiere Werte zu Strings für JSON-Kompatibilität
                    if cell_value is None:
                        row_dict[headers[col_idx]] = ""
                    elif isinstance(cell_value, datetime):
                        row_dict[headers[col_idx]] = cell_value.isoformat()
                    else:
                        row_dict[headers[col_idx]] = str(cell_value)
            
            # Füge Zeile hinzu, wenn sie mindestens einen nicht-leeren Wert hat
            if any(value for value in row_dict.values()):
                rows_data.append(row_dict)
        
        # Wenn keine Datenzeilen gefunden, aber Header vorhanden, erstelle trotzdem eine Tabelle
        if not rows_data and headers:
            # Versuche, auch die Header-Zeile selbst als Daten zu verwenden, wenn sie Daten enthält
            header_row_data = {}
            for col_idx, cell_value in enumerate(header_cells):
                if col_idx < len(headers):
                    if cell_value is None:
                        header_row_data[headers[col_idx]] = ""
                    elif isinstance(cell_value, datetime):
                        header_row_data[headers[col_idx]] = cell_value.isoformat()
                    else:
                        header_row_data[headers[col_idx]] = str(cell_value)
            if any(value for value in header_row_data.values()):
                rows_data.append(header_row_data)
        
        if not rows_data:
            return None
        
        return {
            "sheet_name": sheet_name,
            "header_row": header_row,
            "headers": headers,
            "rows": rows_data,
            "row_count": len(rows_data),
            "column_count": len(headers),
            "quelle": {
                **source_info,
                "blatt": sheet_name
            }
        }
    
    def _extract_unstructured_text(self, sheet, sheet_name: str, source_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrahiert unstrukturierten Text aus Zellen, die nicht Teil einer Tabelle sind"""
        unstructured_texts = []
        
        # Bestimme die Tabellen-Bereiche (Header + Datenzeilen)
        header_row = self._find_header_row(sheet)
        table_rows = set()
        
        if header_row:
            # Alle Zeilen, die Teil einer Tabelle sein könnten
            MAX_ROWS_TO_EXTRACT = 10000
            max_row = min(sheet.max_row, header_row + MAX_ROWS_TO_EXTRACT)
            for row_idx in range(header_row, max_row + 1):
                table_rows.add(row_idx)
        
        # Iteriere über alle Zellen und sammle Text außerhalb von Tabellen
        MAX_CELLS_TO_PROCESS = 50000  # Limit für Performance
        cells_processed = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10000)), start=1):
            # Überspringe Zeilen, die Teil einer Tabelle sind
            if row_idx in table_rows:
                continue
            
            for col_idx, cell in enumerate(row, start=1):
                if cells_processed >= MAX_CELLS_TO_PROCESS:
                    break
                
                cells_processed += 1
                
                # Prüfe, ob Zelle Text enthält
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    
                    # Überspringe leere Zellen oder sehr kurze Werte (wahrscheinlich Formatierung)
                    if not cell_value or len(cell_value) < 2:
                        continue
                    
                    # Überspringe Zahlen ohne Kontext (wahrscheinlich Teil einer Tabelle)
                    try:
                        float(cell_value.replace(',', '.'))
                        # Wenn es eine Zahl ist und wir nicht in einer Tabellen-Zeile sind,
                        # könnte es trotzdem Text sein - behalte es
                    except ValueError:
                        pass  # Keine Zahl, also Text
                    
                    # Füge Text hinzu
                    unstructured_texts.append({
                        "content": cell_value,
                        "sheet": sheet_name,
                        "row": row_idx,
                        "column": col_idx,
                        "cell_reference": f"{self._column_letter(col_idx)}{row_idx}",
                        "quelle": {
                            **source_info,
                            "blatt": sheet_name,
                            "zelle": f"{self._column_letter(col_idx)}{row_idx}"
                        }
                    })
                
                if cells_processed >= MAX_CELLS_TO_PROCESS:
                    break
            
            if cells_processed >= MAX_CELLS_TO_PROCESS:
                break
        
        return unstructured_texts
    
    def _column_letter(self, col_idx: int) -> str:
        """Konvertiert Spaltenindex (1-based) zu Excel-Spaltenbuchstaben (A, B, C, ..., AA, AB, ...)"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + (col_idx % 26)) + result
            col_idx //= 26
        return result
